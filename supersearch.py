#%%
# crawlws through a folder recursively to extract text from common office files
# saves text to csv file
import os, pathlib, csv
from PyPDF2 import PdfReader
import pandas as pd
import textract
from openpyxl import load_workbook
import docx2txt
import logging
import datetime as dt
import multiprocessing.pool
import sqlite3

# get current time
now = dt.datetime.now().strftime("%Y%m%d-%H:%M:%S")
# init global variables
global conn
global c

### CONFIG
db_file = 'fileIndex.sqlite3'
#srcFolder = r'C:\auxDrive\work'
srcFolder = r'/mnt/c/auxDrive/work'
logfile = f'log_{now}.txt'
logging.basicConfig(encoding='utf-8', level=logging.DEBUG)

### END CONFIG
#%%
# initialize db
def dbinit(db_file):
    # creates db connection and creates table
    # if it doesn't exist
    initDbsqlcmd = """
    CREATE TABLE IF NOT EXISTS documents (
	id integer PRIMARY KEY AUTOINCREMENT,
	filepath text NOT NULL,
	content text,
	extension text,
	modifiedDate text
)"""
    conn = None

    try:
        conn = sqlite3.connect(db_file)
        c = conn.cursor()
        c.execute(initDbsqlcmd)
        conn.commit()
    except sqlite3.Error as e:
        logging.error(e)
    if conn:
        return conn
    
# extracts text from pdfs. requires PyPDF2
def extract_text_from_pdf(pdf_file: str):
    # Open the PDF file of your choice
    with open(pdf_file, 'rb') as pdf:
        reader = PdfReader(pdf)
        # no_pages = len(reader.pages)
        pdf_text = ''

        for page in reader.pages:
            content = page.extract_text()
            pdf_text += content

        #pdf_text is an array, concatenate everything

        return pdf_text

def extract_text_general(src: str):
    text = textract.process(src)
    return text

# extracts text from .xlsx and .xls files
# highly unlikely it processes chart text
def extract_text_excel(src: str):
    # iterates over all worksheets and rows/cols to return a text file
    wb = load_workbook(src)
    for worksheet in wb.sheetnames:
        sheet = wb[worksheet]
        max_row = sheet.max_row
        max_col = sheet.max_column

        text = ''
        # iterate over all cells 
        # iterate over all rows
        for i in range(1,max_row+1):
            
            # iterate over all columns
            for j in range(1,max_col+1):          # get particular cell value    
                cell_obj=sheet.cell(row=i,column=j)          # print cell value     
                text += str(cell_obj.value)
                text += ','   # print new line
            text += '\n'
        return text

def extract_text_word(src: str):
    try:
        text = docx2txt.process(src)
    except Exception as e:
        #print(e)
        try:
            text = textract.process(src, method='antiword')
            return text
        except Exception as e:
            #print(e)
            logging.error(e)
            logging.error(f'error while processing {src}, docx2txt and antiword failed to process file')
            return ''
    return text

def processDocument(filepath: str):
    global conn
    global c
    filepath = filepath[0]
    msg = f'processing {filepath}'
    #print(msg)
    logging.info(msg)
    
    # supported list of extensions that we want to process
    # with textract
    textractExtension = ['.xls']

    # get file extension
    print('filepath: ', filepath)
    extension = pathlib.Path(filepath).suffix

    # initializes text variable in case the file is not processed
    text = ''
    # switch to determine what to do with file
    match extension:
        case extension if extension in textractExtension:
            text = extract_text_general(filepath)
        case '.pdf':
            text = extract_text_from_pdf(filepath)
        case '.xlsx':
            text = extract_text_excel(filepath)
        case '.doc' | '.docx':
            text = extract_text_word(filepath)

    # return text and filepath to list
    return [filepath, text, extension]

def main():
    global conn
    global c
    pool = multiprocessing.pool.Pool(processes=8)
     # initialize list. to eventually be added to a db
    tmpList = []

    # initializes db connection
    conn = dbinit(db_file)
    c = conn.cursor()

    # load csv with file index
    with open('fileIndex.txt', 'r') as f:
        reader = csv.reader(f)
        fileList = list(reader)
    
    #print(fileList)
    for file in fileList:
        processedData = processDocument(file)
        tmpList.append(processedData)
        c.execute('INSERT INTO documents (filepath, content, extension) VALUES (?,?,?);', (processedData[0], processedData[1], processedData[2]))
        conn.commit()
    #tmpList = pool.map(processDocument,fileList, chunksize=3)
    # put data into df
    msg = 'saving data to df'
    print(msg)
    logging.info(msg)
    df = pd.DataFrame(tmpList, columns=['filepath', 'content', 'extension'])
    conn.close()

    return df

# extracts text from docx files
#extract_text_from_pdf("/home/richard/Documents/supersearch/samplefolderstructure/CWEST Enhancements/PL-0001 OX Kicker Stroke Investigation/K-402191-PL-0001 R00.pdf")
#text = textract.process("/home/richard/Documents/supersearch/samplefolderstructure/CWEST Enhancements/PL-0001 OX Kicker Stroke Investigation/K-402191-PL-0001 R00a.docx")
#print(text)
df = main()